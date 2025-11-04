import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import re
import requests

st.set_page_config(page_title="Google Sheet → Excel 변환기", layout="wide")


# ✅ 캐시 완전 비활성화 (자동 새로고침 보장)
def nocache():
    return None


# ✅ 구글시트 CSV 불러오기 (캐시 X)
def load_sheet_csv(sheet_url):
    try:
        csv_url = sheet_url.replace("/edit?usp=sharing", "").replace("/edit", "") + "/export?format=csv"
        df = pd.read_csv(csv_url, header=None)
        return df
    except Exception as e:
        st.error(f"❌ 구글 시트를 불러오는 중 오류 발생: {e}")
        return None


# ✅ 전화번호 정규화 함수
def normalize_phone(phone):
    if pd.isna(phone):
        return ""
    phone = str(phone).replace("-", "").replace(" ", "").replace("+82", "0")
    if phone.startswith("82") and len(phone) >= 11:
        phone = "0" + phone[2:]
    if len(phone) == 10:
        phone = "0" + phone
    if len(phone) == 11:
        return f"{phone[0:3]}-{phone[3:7]}-{phone[7:11]}"
    return phone


# ✅ 엑셀 변환 함수
def convert_to_excel(df):
    today = datetime.now().strftime("%Y%m%d")
    df = df.copy()

    # ◾ 첫 번째 열 결측 채우기 (연월일 + 2자리 순번)
    count = len(df)
    fill_values = [f"{today}{num:02d}" for num in range(1, count + 1)]
    df.iloc[:, 0] = fill_values

    # ◾ 전화번호 정규화 (6번째 열 = index 5)
    if df.shape[1] >= 6:
        df.iloc[:, 5] = df.iloc[:, 5].apply(normalize_phone)

    # ◾ 엑셀 저장 + 서식 적용
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # 1️⃣ 첫 행 삭제
    ws.delete_rows(1)

    # 2️⃣ 테두리 적용
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # 3️⃣ 열너비 자동 맞춤
    def visual_len(s: str) -> int:
        if s is None:
            return 0
        s = str(s)
        length = 0
        for ch in s:
            length += 1 if ord(ch) <= 255 else 2
        return length

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, visual_len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    # 4️⃣ 파일 저장 후 반환
    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    return output_final, f"filled_sheet_{today}.xlsx"


# ============= Streamlit UI =============

st.title("📄 Google Sheet → Excel 자동 변환기")
st.write("✅ 결측 데이터 자동 채움 · ✅ 전화번호 자동 정규화 · ✅ 테두리/열너비 자동 적용")

sheet_url = st.text_input("📌 Google Sheet URL 입력", placeholder="https://docs.google.com/spreadsheets/d/XXXXX/edit?usp=sharing")

if sheet_url:
    if st.button("🔄 최신 데이터 불러오기 (캐시 제거)"):
        st.toast("⏳ 데이터를 불러오는 중...", icon="⏳")
        df = load_sheet_csv(sheet_url)

        if df is not None:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.info(f"📌 데이터 불러온 시각: **{now}**")

            # 첫 번째 열이 비었거나 NaN인 행만 변환 대상
            df_missing = df[df.iloc[:, 0].isna() | (df.iloc[:, 0] == "")]
            row_count = len(df_missing)

            st.success(f"📊 변환 대상 행 수: **{row_count} rows**")

            if row_count == 0:
                st.warning("⚠️ 변환할 대상이 없습니다. (이미 모든 행에 값이 있음)")
            else:
                excel_binary, excel_name = convert_to_excel(df_missing)

                st.download_button(
                    label="⬇️ 엑셀 다운로드",
                    data=excel_binary,
                    file_name=excel_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.toast("✅ 변환 완료! 다운로드 시작됨", icon="✅")

else:
    st.warning("👆 먼저 Google Sheet URL을 입력해주세요!")
